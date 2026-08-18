import json
from datetime import date
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import MOSTRAR_PO_UI

try:
    from src import db as _db
except ImportError:
    _db = None

_COLUNAS_PO = ["IA+PO — combinação", "IA+PO — como usar", "IA+PO — exemplo", "IA+PO — risco"]

_JSON_PATH = Path(__file__).resolve().parent.parent / "data" / "pmbok_processos.json"
_MATRIZ_PATH = Path(__file__).resolve().parent.parent / "data" / "matriz_pilotos_processos.json"
_PILOTOS_PATH = Path(__file__).resolve().parent.parent / "data" / "pilotos.json"
_CASES_PATH = Path(__file__).resolve().parent.parent / "data" / "cases_bezerra.json"


def _carregar_cases_index() -> dict:
    if not _CASES_PATH.exists():
        return {}
    with open(_CASES_PATH, encoding="utf-8") as f:
        cases = json.load(f).get("cases", [])
    return {c["id"]: c for c in cases}

_CORES_AREA = {
    "Governança": "2E7D32",
    "Escopo": "F9A825",
    "Cronograma": "1565C0",
    "Finanças": "EF6C00",
    "Partes Interessadas": "757575",
    "Recursos": "5D4037",
    "Riscos": "C62828",
}

_ORDEM_GRUPOS = [
    "Início",
    "Planejamento",
    "Execução",
    "Monitoramento e Controle",
    "Encerramento",
]


def carregar_processos() -> list[dict]:
    if not _JSON_PATH.exists():
        return []
    with open(_JSON_PATH, encoding="utf-8") as f:
        data = json.load(f)
    return data.get("processos", [])


def _dataframe(processos: list[dict]) -> pd.DataFrame:
    linhas = []
    for p in processos:
        linhas.append(
            {
                "ID": p["id"],
                "Área": p["area"],
                "Grupo": p["grupo"],
                "Processo": p["nome"],
                "IA — ferramenta": p["ia"].get("ferramenta", ""),
                "IA — como usar": p["ia"].get("como_usar", ""),
                "IA — exemplo": p["ia"].get("exemplo", ""),
                "IA — risco": p["ia"].get("risco", ""),
                "IA+PO — combinação": p["ia_po"].get("combinacao", ""),
                "IA+PO — como usar": p["ia_po"].get("como_usar", ""),
                "IA+PO — exemplo": p["ia_po"].get("exemplo", ""),
                "IA+PO — risco": p["ia_po"].get("risco", ""),
            }
        )
    return pd.DataFrame(linhas)


def exportar_xlsx(processos: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "PMBOK x IA x PO"

    header = [
        "ID", "Área", "Grupo", "Processo",
        "IA — ferramenta", "IA — como usar", "IA — exemplo", "IA — risco",
        "IA+PO — combinação", "IA+PO — como usar", "IA+PO — exemplo", "IA+PO — risco",
    ]

    for col_idx, titulo in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="263238", end_color="263238", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, p in enumerate(processos, start=2):
        cor = _CORES_AREA.get(p["area"], "BDBDBD")
        valores = [
            p["id"], p["area"], p["grupo"], p["nome"],
            p["ia"].get("ferramenta", ""), p["ia"].get("como_usar", ""),
            p["ia"].get("exemplo", ""), p["ia"].get("risco", ""),
            p["ia_po"].get("combinacao", ""), p["ia_po"].get("como_usar", ""),
            p["ia_po"].get("exemplo", ""), p["ia_po"].get("risco", ""),
        ]
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 2:
                cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

    larguras = [6, 16, 24, 34, 22, 44, 44, 34, 22, 44, 44, 34]
    for idx, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "E2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_markdown_obsidian(processos: list[dict]) -> str:
    hoje = date.today().isoformat()
    linhas = [
        "---",
        "created: " + hoje,
        "updated: " + hoje,
        "area: pesquisa",
        "type: referencia",
        "status: ativo",
        "tags: [pesquisa/referencia, status/ativo]",
        "aliases: [\"Mapa PMBOK 8 IA PO\", \"PMBOK IA+PO\"]",
        "source: PMBOK 8a Edicao — mapa BSBr",
        "---",
        "",
        "# Mapa PMBOK 8ª Ed. × IA × IA+PO",
        "",
        "Referência de uso de IA (só-IA) e IA combinada com Pesquisa Operacional (IA+PO)",
        "para cada um dos 40 processos do PMBOK 8ª Edição. Alimenta o app [[consultor-ia-pppm]]",
        "(porta 8512) e serve como base para consultoria PME e material didático.",
        "",
        "Relacionado: [[_MOC Pesquisa]] · [[MBA em Pesquisa Operacional]] · [[Framework Eixo Estratégico]]",
        "",
        "## Sumário por área",
        "",
        "| Área | Nº processos | Grupos que atravessa |",
        "|------|-------------|----------------------|",
    ]

    areas_ordenadas = list({p["area"] for p in processos})
    ordem_areas = [
        "Governança", "Escopo", "Cronograma", "Finanças",
        "Partes Interessadas", "Recursos", "Riscos",
    ]
    areas_ordenadas.sort(key=lambda a: ordem_areas.index(a) if a in ordem_areas else 999)

    for area in areas_ordenadas:
        procs_area = [p for p in processos if p["area"] == area]
        grupos = sorted({p["grupo"] for p in procs_area}, key=lambda g: _ORDEM_GRUPOS.index(g) if g in _ORDEM_GRUPOS else 999)
        linhas.append(f"| {area} | {len(procs_area)} | {', '.join(grupos)} |")

    linhas.append("")

    for area in areas_ordenadas:
        linhas.append(f"## {area}")
        linhas.append("")
        procs_area = sorted(
            [p for p in processos if p["area"] == area], key=lambda p: p["id"]
        )
        for p in procs_area:
            linhas.append(f"### {p['id']} — {p['nome']}")
            linhas.append(f"*Grupo: {p['grupo']}*")
            linhas.append("")
            linhas.append("**Só IA**")
            linhas.append(f"- **Ferramenta:** {p['ia'].get('ferramenta', '')}")
            linhas.append(f"- **Como usar:** {p['ia'].get('como_usar', '')}")
            linhas.append(f"- **Exemplo:** {p['ia'].get('exemplo', '')}")
            linhas.append(f"- **Risco:** {p['ia'].get('risco', '')}")
            linhas.append("")
            linhas.append("**IA + PO**")
            linhas.append(f"- **Combinação:** {p['ia_po'].get('combinacao', '')}")
            linhas.append(f"- **Como usar:** {p['ia_po'].get('como_usar', '')}")
            linhas.append(f"- **Exemplo:** {p['ia_po'].get('exemplo', '')}")
            linhas.append(f"- **Risco:** {p['ia_po'].get('risco', '')}")
            linhas.append("")

    linhas.append("---")
    linhas.append("*Gerado por `consultor-ia-pppm` — módulo `mapa_pmbok`*")
    return "\n".join(linhas)


def _carregar_matriz() -> dict:
    if not _MATRIZ_PATH.exists():
        return {}
    with open(_MATRIZ_PATH, encoding="utf-8") as f:
        return json.load(f).get("relacoes", {})


def _carregar_pilotos() -> list[dict]:
    if not _PILOTOS_PATH.exists():
        return []
    with open(_PILOTOS_PATH, encoding="utf-8") as f:
        return json.load(f)


def _renderizar_matriz_cruzada(processos: list[dict]) -> None:
    st.markdown("### 🔀 Matriz cruzada: processo PMBOK × piloto de IA")
    st.caption(
        "Para cada processo, quais dos 12 pilotos do catálogo o endereçam. "
        "🟩 = piloto primário (resposta principal). 🟦 = piloto secundário (contribui parcialmente)."
    )
    relacoes = _carregar_matriz()
    pilotos = _carregar_pilotos()
    if not relacoes or not pilotos:
        st.warning(
            "Matriz ou catálogo de pilotos ausente. Confira "
            "`data/matriz_pilotos_processos.json` e `data/pilotos.json`."
        )
        return

    # Índice processo -> lista de (piloto_nome, papel)
    idx: dict[str, list[tuple[str, str]]] = {}
    id_para_nome = {p["id"]: p["nome"] for p in pilotos}
    for piloto_id, papeis in relacoes.items():
        nome = id_para_nome.get(piloto_id, piloto_id)
        for pid in papeis.get("primaria", []):
            idx.setdefault(pid, []).append((nome, "primaria"))
        for pid in papeis.get("secundaria", []):
            idx.setdefault(pid, []).append((nome, "secundaria"))

    linhas = []
    for p in processos:
        aderentes = idx.get(p["id"], [])
        prim = [n for n, r in aderentes if r == "primaria"]
        secn = [n for n, r in aderentes if r == "secundaria"]
        linhas.append(
            {
                "Proc.": p["id"],
                "Área": p["area"],
                "Nome do processo": p["nome"],
                "Pilotos primários (🟩)": " · ".join(prim) if prim else "—",
                "Pilotos secundários (🟦)": " · ".join(secn) if secn else "—",
                "Cobertura": len(prim) + len(secn),
            }
        )

    df_mat = pd.DataFrame(linhas)
    st.dataframe(
        df_mat,
        use_container_width=True,
        height=520,
        column_config={
            "Nome do processo": st.column_config.TextColumn(width="large"),
            "Pilotos primários (🟩)": st.column_config.TextColumn(width="large"),
            "Pilotos secundários (🟦)": st.column_config.TextColumn(width="large"),
            "Cobertura": st.column_config.NumberColumn(help="Nº de pilotos que endereçam o processo (primário + secundário)"),
        },
    )

    total_procs = len(processos)
    procs_cobertos = sum(1 for l in linhas if l["Cobertura"] > 0)
    procs_sem_piloto = total_procs - procs_cobertos
    c1, c2, c3 = st.columns(3)
    c1.metric("Processos cobertos", f"{procs_cobertos}/{total_procs}")
    c2.metric("Processos sem piloto", procs_sem_piloto, help="Gap explícito de catálogo — oportunidade para novo piloto")
    c3.metric("Pilotos no catálogo", len(pilotos))


_DIMENSOES_ALVO = {
    "estrategia": "Estratégia",
    "dados": "Dados",
    "casos_uso": "Casos de uso",
    "governanca": "Governança",
    "beneficios": "Benefícios",
}


def _renderizar_pilotos_didatico(pilotos: list[dict]) -> None:
    """Visão principal da aba 1: os 16 pilotos como cards didáticos."""
    idx_cases = _carregar_cases_index()
    citados = [p for p in pilotos if p.get("citado_por_bezerra")]
    complementares = [p for p in pilotos if not p.get("citado_por_bezerra")]

    st.markdown(
        "Foco prático: **onde IA muda o jogo em Portfólio, Programa e Projeto (PPPM)**. "
        "Cada piloto abaixo é uma aplicação concreta com dor real, exemplo, "
        "esforço estimado, ganho esperado e KPI benchmark auditável. "
        "Os 🎯 foram citados pelo Prof. Bezerra na Aula 1."
    )

    with st.container(border=True):
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total de pilotos", len(pilotos))
        c2.metric("Citados por Bezerra", len(citados))
        c3.metric("Complementares", len(complementares))
        tempo_medio = sum(p.get("tempo_estimado_semanas", 0) for p in pilotos) / max(len(pilotos), 1)
        c4.metric("Tempo médio (sem.)", f"{tempo_medio:.1f}")

    dims = ["(todas)"] + list(_DIMENSOES_ALVO.values())
    filtro_dim = st.selectbox(
        "Filtrar por dimensão do diagnóstico",
        dims,
        key="pilotos_didatico_dim",
        help="Dimensões da aba 3 (Diagnóstico): Estratégia, Dados, Casos de uso, Governança, Benefícios.",
    )
    filtro_bezerra = st.checkbox(
        "Mostrar só os 🎯 citados pelo Prof. Bezerra",
        value=False,
        key="pilotos_didatico_bezerra",
    )
    busca = st.text_input(
        "Buscar termo no piloto (nome, descrição, ferramentas)", "",
        key="pilotos_didatico_busca",
    )

    def _match(p: dict) -> bool:
        if filtro_bezerra and not p.get("citado_por_bezerra"):
            return False
        if filtro_dim != "(todas)":
            rot_para_id = {v: k for k, v in _DIMENSOES_ALVO.items()}
            if rot_para_id[filtro_dim] not in (p.get("dimensoes_alvo") or []):
                return False
        if busca.strip():
            alvo = " ".join([
                p.get("nome", ""), p.get("descricao", ""),
                " ".join(p.get("ferramentas_recomendadas") or []),
                " ".join(p.get("categorias_dor") or []),
            ]).lower()
            if busca.lower() not in alvo:
                return False
        return True

    pilotos_view = [p for p in pilotos if _match(p)]
    st.caption(f"{len(pilotos_view)} de {len(pilotos)} pilotos após filtros.")

    for p in pilotos_view:
        emoji = " 🎯" if p.get("citado_por_bezerra") else ""
        titulo = f"**{p['nome']}**{emoji}"
        with st.expander(titulo, expanded=False):
            st.markdown(p.get("descricao", ""))

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Impacto", str(p.get("impacto_base", "—")).capitalize())
            c2.metric("Viabilidade", str(p.get("viabilidade_base", "—")).capitalize())
            c3.metric("Risco", str(p.get("risco_base", "—")).capitalize())
            c4.metric("Tempo (sem.)", p.get("tempo_estimado_semanas", "—"))

            col_a, col_b = st.columns(2)
            with col_a:
                st.markdown("**🎯 Dimensões que endereça**")
                for d in p.get("dimensoes_alvo") or []:
                    st.markdown(f"- {_DIMENSOES_ALVO.get(d, d)}")
                st.markdown("**🩹 Categorias de dor**")
                for c in p.get("categorias_dor") or []:
                    st.markdown(f"- {c}")
            with col_b:
                st.markdown("**⚙️ Ferramentas recomendadas**")
                for f in p.get("ferramentas_recomendadas") or []:
                    st.markdown(f"- {f}")
                st.markdown("**📋 Pré-requisitos**")
                for pr in p.get("pre_requisitos") or []:
                    st.markdown(f"- {pr}")

            st.success(f"**Ganho esperado:** {p.get('ganho_esperado', '—')}")

            if p.get("metricas"):
                st.markdown("**📊 KPI benchmark auditável**")
                for m in p["metricas"]:
                    with st.container(border=True):
                        st.markdown(f"**{m.get('nome', 'KPI')}** — unidade: `{m.get('unidade', '')}`")
                        base = m.get("baseline_mercado", {})
                        meta = m.get("meta_com_ia", {})
                        st.markdown(
                            f"- Baseline mercado: {base.get('valor', '—')}  \n"
                            f"  *Fonte: {base.get('fonte', '—')}*"
                        )
                        st.markdown(
                            f"- Meta com IA: {meta.get('valor', '—')}  \n"
                            f"  *Fonte: {meta.get('fonte', '—')}*"
                        )
                        if m.get("reducao_percentual"):
                            st.markdown(f"- 🏁 **Ganho:** {m['reducao_percentual']}")

            if p.get("plano_projeto_30d"):
                with st.expander("📅 Plano de 30 dias (EAP + cronograma)"):
                    plano = p["plano_projeto_30d"]
                    if isinstance(plano, dict):
                        for k, v in plano.items():
                            st.markdown(f"**{k}:** {v}" if not isinstance(v, list)
                                        else f"**{k}:**\n" + "\n".join(f"- {x}" for x in v))
                    else:
                        st.write(plano)


def render() -> None:
    titulo_extra = " × IA+PO" if MOSTRAR_PO_UI else ""
    caption_extra = " × IA + Pesquisa Operacional" if MOSTRAR_PO_UI else ""
    st.subheader("Pilotos de IA em PPPM · foco prático")
    st.caption(
        "Onde a IA muda o jogo em Portfólio, Programa e Projeto. "
        f"16 pilotos com dor, exemplo, esforço e KPI. "
        f"O mapa completo dos 40 processos PMBOK{titulo_extra} fica como referência opcional no fim da aba{caption_extra}."
    )

    pilotos = _carregar_pilotos()
    if pilotos:
        _renderizar_pilotos_didatico(pilotos)

    st.markdown("---")
    with st.expander("📚 Referência opcional — Mapa completo dos 40 processos PMBOK 8ª Ed.", expanded=False):
        _render_mapa_40_processos()


def _render_mapa_40_processos() -> None:

    processos = carregar_processos()
    if not processos:
        st.error(
            "Arquivo `data/pmbok_processos.json` não encontrado ou vazio. "
            "Rode o gerador de dataset primeiro."
        )
        return

    df = _dataframe(processos)

    projeto_ativo = None
    tratamentos_ativo: dict = {}
    if _db is not None and st.session_state.get("projeto_ativo_id"):
        projeto_ativo = _db.obter_projeto(st.session_state["projeto_ativo_id"])
        if projeto_ativo:
            tratamentos_ativo = _db.listar_tratamentos_pmbok(projeto_ativo["id"])
            st.info(
                f"🎯 Projeto ativo: **{projeto_ativo['nome']}** — "
                f"{len(tratamentos_ativo)} processos marcados. "
                "Marcações aparecem com badge nas fichas."
            )

    with st.container(border=True):
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Processos", len(processos))
        c2.metric("Áreas", df["Área"].nunique())
        c3.metric("Grupos", df["Grupo"].nunique())
        c4.metric("Marcados no ativo", len(tratamentos_ativo) if projeto_ativo else 0)

    col_f1, col_f2, col_f3, col_f4 = st.columns([2, 2, 2, 3])
    areas = ["(todas)"] + sorted(df["Área"].unique().tolist())
    grupos = ["(todos)"] + [g for g in _ORDEM_GRUPOS if g in df["Grupo"].unique().tolist()]
    filtro_area = col_f1.selectbox("Filtrar por área", areas, key="mapa_pmbok_area")
    filtro_grupo = col_f2.selectbox("Filtrar por grupo", grupos, key="mapa_pmbok_grupo")
    filtro_ativo = col_f3.selectbox(
        "Filtro projeto ativo",
        ["(todos)", "Só marcados", "Só não marcados", "Só gaps", "Só críticos"],
        help="Filtra pelas marcações do projeto ativo (definido na aba 7).",
        key="mapa_pmbok_filtro_ativo",
    )
    busca = col_f4.text_input("Buscar termo em qualquer campo", "", key="mapa_pmbok_busca")

    df_view = df.copy()
    if filtro_area != "(todas)":
        df_view = df_view[df_view["Área"] == filtro_area]
    if filtro_grupo != "(todos)":
        df_view = df_view[df_view["Grupo"] == filtro_grupo]
    if busca.strip():
        mask = df_view.apply(
            lambda r: r.astype(str).str.contains(busca, case=False, na=False).any(), axis=1
        )
        df_view = df_view[mask]

    if projeto_ativo and filtro_ativo != "(todos)":
        if filtro_ativo == "Só marcados":
            df_view = df_view[df_view["ID"].isin(tratamentos_ativo.keys())]
        elif filtro_ativo == "Só não marcados":
            df_view = df_view[~df_view["ID"].isin(tratamentos_ativo.keys())]
        elif filtro_ativo == "Só gaps":
            gap_ids = [k for k, v in tratamentos_ativo.items() if v["tratamento"] == "gap"]
            df_view = df_view[df_view["ID"].isin(gap_ids)]
        elif filtro_ativo == "Só críticos":
            crit_ids = [k for k, v in tratamentos_ativo.items() if v["criticidade"] == "alta"]
            df_view = df_view[df_view["ID"].isin(crit_ids)]

    if projeto_ativo and tratamentos_ativo:
        def _ord_criticidade(pid: str) -> int:
            t = tratamentos_ativo.get(pid)
            if not t:
                return 3
            return {"alta": 0, "media": 1, "baixa": 2}.get(t["criticidade"], 3)
        df_view = df_view.assign(_ord=df_view["ID"].map(_ord_criticidade)).sort_values(
            ["_ord", "ID"]
        ).drop(columns=["_ord"])

    modo = st.radio(
        "Modo de visualização",
        ["Fichas expansíveis", "Tabela comparativa", "Matriz cruzada (processo × piloto)"],
        horizontal=True,
        key="mapa_pmbok_modo",
    )

    if modo == "Matriz cruzada (processo × piloto)":
        _renderizar_matriz_cruzada(processos)
        return

    processos_view = [p for p in processos if p["id"] in df_view["ID"].tolist()]

    if modo == "Tabela comparativa":
        df_render = df_view if MOSTRAR_PO_UI else df_view.drop(columns=_COLUNAS_PO, errors="ignore")
        col_cfg = {
            "IA — como usar": st.column_config.TextColumn(width="large"),
            "IA — exemplo": st.column_config.TextColumn(width="large"),
        }
        if MOSTRAR_PO_UI:
            col_cfg["IA+PO — como usar"] = st.column_config.TextColumn(width="large")
            col_cfg["IA+PO — exemplo"] = st.column_config.TextColumn(width="large")
        st.dataframe(df_render, use_container_width=True, height=520, column_config=col_cfg)
    else:
        for p in processos_view:
            cor = _CORES_AREA.get(p["area"], "607D8B")
            marc = tratamentos_ativo.get(p["id"])
            badge_emoji = ""
            if marc:
                badge_emoji = {
                    "ia": " 🤖", "ia_po": " 🧮", "gap": " ⚠️", "nenhum": " ⬜",
                }.get(marc["tratamento"], "")
                if marc["criticidade"] == "alta":
                    badge_emoji += " 🔴"
            titulo = f"**{p['id']} — {p['nome']}**{badge_emoji} · {p['area']} · {p['grupo']}"
            with st.expander(titulo):
                if marc:
                    st.markdown(
                        f"<div style='background:#E8F5E9;padding:8px;border-left:3px solid #43A047;"
                        f"border-radius:3px;margin-bottom:8px'>"
                        f"<b>Marcado por {projeto_ativo['nome']}:</b> "
                        f"tratamento <code>{marc['tratamento']}</code> · "
                        f"criticidade <code>{marc['criticidade']}</code>"
                        f"{'<br><i>' + marc['observacao'] + '</i>' if marc['observacao'] else ''}"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                st.markdown(
                    f"<div style='background:#{cor};color:white;padding:4px 10px;"
                    f"border-radius:4px;display:inline-block;font-size:12px'>{p['area']}</div>",
                    unsafe_allow_html=True,
                )
                if MOSTRAR_PO_UI:
                    col_ia, col_iapo = st.columns(2)
                else:
                    col_ia = st.container()
                    col_iapo = None
                with col_ia:
                    st.markdown("### 🤖 Aplicação de IA")
                    st.markdown(f"**Ferramenta:** {p['ia'].get('ferramenta', '')}")
                    st.markdown(f"**Como usar:** {p['ia'].get('como_usar', '')}")
                    st.markdown(f"**Exemplo:** {p['ia'].get('exemplo', '')}")
                    st.markdown(
                        f"<div style='background:#FFF3E0;padding:8px;border-left:3px solid #FB8C00;"
                        f"border-radius:3px'><b>⚠ Risco:</b> {p['ia'].get('risco', '')}</div>",
                        unsafe_allow_html=True,
                    )
                if MOSTRAR_PO_UI and col_iapo is not None:
                    with col_iapo:
                        st.markdown("### 🧮 IA + PO")
                        st.markdown(f"**Combinação:** {p['ia_po'].get('combinacao', '')}")
                        st.markdown(f"**Como usar:** {p['ia_po'].get('como_usar', '')}")
                        st.markdown(f"**Exemplo:** {p['ia_po'].get('exemplo', '')}")
                        st.markdown(
                            f"<div style='background:#FFF3E0;padding:8px;border-left:3px solid #FB8C00;"
                            f"border-radius:3px'><b>⚠ Risco:</b> {p['ia_po'].get('risco', '')}</div>",
                            unsafe_allow_html=True,
                        )

                # Métricas de mercado (benchmark) — opcional
                if p.get("metricas"):
                    st.markdown("### 📊 Métricas de mercado (benchmark auditável)")
                    for m in p["metricas"]:
                        st.markdown(f"**{m.get('nome','KPI')}** · unidade: `{m.get('unidade','')}`")
                        base = m.get("baseline_mercado", {})
                        meta = m.get("meta_com_ia", {})
                        st.markdown(
                            f"- Baseline: {base.get('valor','—')} · fonte: {base.get('fonte','—')} · "
                            f"confiança: `{base.get('confianca','—')}`"
                        )
                        st.markdown(
                            f"- Meta com IA: {meta.get('valor','—')} · fonte: {meta.get('fonte','—')} · "
                            f"confiança: `{meta.get('confianca','—')}`"
                        )
                        if m.get("reducao_percentual"):
                            st.markdown(f"- Ganho: **{m['reducao_percentual']}**")

                # Formatos de entrada de dados aceitos — opcional
                if p.get("formatos_entrada_dados"):
                    st.markdown("### 📥 Formatos de entrada aceitos")
                    for f in p["formatos_entrada_dados"]:
                        st.markdown(f"- **{f.get('tipo','?')}** — {f.get('descricao','')}")
                        if f.get("exemplo_estrutura"):
                            st.caption(f"  Estrutura típica: {f['exemplo_estrutura']}")

                # Casos reais do Bezerra aplicáveis — opcional
                if p.get("casos_bezerra"):
                    st.markdown("### 📋 Casos reais aplicáveis (Bezerra, Aula 1)")
                    idx_cases = _carregar_cases_index()
                    for cid in p["casos_bezerra"]:
                        case = idx_cases.get(cid)
                        if case:
                            st.markdown(
                                f"- **{case.get('setor','')} · {case.get('porte','')}** — "
                                f"{case.get('resultado_numerico','')}"
                            )
                            st.caption(f"  {case.get('citacao_bezerra','')}")

    st.markdown("---")
    st.markdown("### 📥 Exportar")
    col_x, col_m = st.columns(2)
    with col_x:
        st.download_button(
            "📊 Baixar .xlsx",
            data=exportar_xlsx(processos_view),
            file_name=f"pmbok_ia_po_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col_m:
        st.download_button(
            "📝 Baixar .md (Obsidian)",
            data=exportar_markdown_obsidian(processos_view).encode("utf-8"),
            file_name=f"pmbok_ia_po_{date.today().isoformat()}.md",
            mime="text/markdown",
            use_container_width=True,
        )
